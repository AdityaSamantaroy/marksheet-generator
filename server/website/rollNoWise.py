import pandas as pd
import os
import shutil
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
# from "./validate.py" import validate
import openpyxl
from validate import check

BASE_PATH = os.path.dirname(
    os.path.realpath(__file__))

OUTPUT_FILE_PATH = BASE_PATH + "/output/rollNumberWise"
RESPONSE_FILE = BASE_PATH + "/uploads/csv/responses.csv"
MASTER_FILE = BASE_PATH + "/uploads/csv/master.csv"
TEMPLATE_FILE = BASE_PATH + "/templates/marksheet_template.xlsx"

# num_questions = 0

stud_info = {}
answers = {}
ans_idx = check()
coords = {
    "name": "B6",
    "rollnumber": "B7",
    "#_right": "B10",
    "#_wrong": "C10",
    "#_left": "D10",
    "pos": "B11",
    "neg": "C11",
    "max": "E10",
    "total": "B12",
    "wrong": "C12",
    "final": "E12"
}

# print(check())


def process_master():
    # pass
    df = pd.read_csv(MASTER_FILE)
    for i in range(len(df)):
        rollno = df.loc[i, "roll"]
        name = df.loc[i, "name"]
        stud_info[rollno] = {
            "name": name,
            "done": False
        }


def get_answer():
    df = pd.read_csv(RESPONSE_FILE)
    print(df.columns)
    num_questions = len(df.columns) - 7
    return num_questions


def generate_marksheet():
    num_questions = get_answer()
    start = 7
    df = pd.read_csv(RESPONSE_FILE)
    red = Font(color="FF5733")
    blue = Font(color="3080FF")
    green = Font(color="6CFF30")
    for i in range(len(df)):
        rollno = df.loc[i, "Roll Number"]
        if rollno == "ANSWER":
            continue
        filename = rollno.upper() + ".xlsx"
        target_path = OUTPUT_FILE_PATH + "/" + filename

        shutil.copyfile(TEMPLATE_FILE, target_path)

        wb = openpyxl.load_workbook(target_path)
        ws = wb['Sheet']
        from views import negative, positive

        correct = 0
        wrong = 0
        left = 0
        begin = 16
        for j in range(num_questions):
            studCellNo = "A"+str(begin+j)
            ansCellNo = "B"+str(begin+j)
            ws[studCellNo] = df.loc[i, "Unnamed "+str(start+j)]
            ws[ansCellNo] = df.loc[ans_idx, "Unnamed" + str(start+j)]
            ws[ansCellNo].font = blue

            if ws[studCellNo] == ws[ansCellNo]:
                correct += 1
                ws[studCellNo].font = green
            elif len(ws[studCellNo]) == 0:
                left += 1
            else:
                wrong += 1
                ws[studCellNo].font = red
        ws.title = 'quiz'

        ws.save(target_path)
        return


# get_answer()
process_master()
get_answer()
generate_marksheet()
